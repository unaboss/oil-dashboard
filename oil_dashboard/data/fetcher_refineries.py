"""Fetch US refinery capacity (annual EIA report) and utilization (weekly API).

Capacity: parsed from the EIA Refinery Capacity Report XLSX (one row per
refinery, crude-distillation capacity in bbl/calendar day). Static reference,
updated annually.

Utilization: live weekly series from the EIA API (percent utilization, crude
inputs, operable capacity) at US and PADD level.
"""

import io
from datetime import datetime, timezone

import requests

from data.cache import get, set
from config import (
    EIA_API_KEY, EIA_BASE,
    REFINERY_CAP_URL, REFINERY_CAP_PRODUCT, REFINERY_CAP_SERIES,
    REFINERY_UTIL_ROUTE, REFINERY_UTIL_SERIES,
    CACHE_TTL_REFINERY_CAP, CACHE_TTL_REFINERY_CAP_FAILURE,
    CACHE_TTL_REFINERY_UTIL,
)

CACHE_KEY_CAP = "refinery_capacity"
CACHE_KEY_UTIL = "refinery_utilization"

TIMEOUT_SECONDS = 30


def get_refinery_capacity(year=None):
    """Return a list of refineries with crude-distillation capacity (bbl/day).

    Each row: {company, site, state, district, padd, capacity_bpd}
    Empty list on any failure (missing file, parse error, network).
    """
    data, _, _, stale = get(CACHE_KEY_CAP)
    if data is not None and not stale:
        return data

    year = year or datetime.now(timezone.utc).year
    rows = _parse_capacity_workbook(REFINERY_CAP_URL.format(yy=str(year)[-2:]))
    # Cache a successful fetch for the long TTL; a failed (empty) fetch is
    # cached briefly so a transient outage retries soon instead of blanking
    # the table for 30 days.
    ttl = CACHE_TTL_REFINERY_CAP if rows else CACHE_TTL_REFINERY_CAP_FAILURE
    set(CACHE_KEY_CAP, rows, ttl,
        last_updated=datetime.now(timezone.utc).isoformat())
    return rows


def _parse_capacity_workbook(url):
    try:
        resp = requests.get(url, timeout=TIMEOUT_SECONDS)
        resp.raise_for_status()
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        return []

    rows = []
    for row in ws.iter_rows(values_only=True):
        if len(row) < 11:
            continue
        product = row[8]
        series = row[9]
        if product != REFINERY_CAP_PRODUCT or series != REFINERY_CAP_SERIES:
            continue
        try:
            capacity = int(row[10])
        except (TypeError, ValueError):
            continue
        rows.append({
            "company": str(row[3] or "").strip(),
            "site": str(row[6] or "").strip(),
            "state": str(row[5] or "").strip(),
            "district": str(row[4] or "").strip(),
            "padd": str(row[7] or "").strip(),
            "capacity_bpd": capacity,
        })

    rows.sort(key=lambda r: r["capacity_bpd"], reverse=True)
    return rows


def get_refinery_utilization(weeks=12):
    """Return weekly utilization series for the US and each PADD.

    {"dates": [iso dates], "utilization_pct": {area: [values]},
     "crude_inputs": {area: [values]}, "operable_capacity": {area: [values]},
     "available": bool}
    """
    data, _, _, stale = get(CACHE_KEY_UTIL)
    if data is not None and not stale:
        return data

    result = {"dates": [], "utilization_pct": {}, "crude_inputs": {},
              "operable_capacity": {}, "available": False}

    areas = ["NUS", "R10", "R20", "R30", "R40", "R50"]
    series_areas = {
        REFINERY_UTIL_SERIES["utilization_pct"]: [],
        REFINERY_UTIL_SERIES["crude_inputs"]: [],
        REFINERY_UTIL_SERIES["operable_capacity"]: [],
    }

    for area in areas:
        params = {
            "api_key": EIA_API_KEY,
            "frequency": "weekly",
            "data[0]": "value",
            "sort[0][column]": "period",
            "sort[0][direction]": "desc",
            "length": weeks,
            "facets[duoarea][]": area,
        }
        try:
            resp = requests.get(
                f"{EIA_BASE}{REFINERY_UTIL_ROUTE}/data", params=params, timeout=TIMEOUT_SECONDS)
            resp.raise_for_status()
            payload = resp.json()
            records = payload.get("response", {}).get("data", [])
        except Exception:
            continue

        by_period = {}
        for rec in records:
            period = rec.get("period")
            series = rec.get("series")
            value = rec.get("value")
            if not period or not series or value is None:
                continue
            try:
                by_period.setdefault(period, {})[series] = float(value)
            except (TypeError, ValueError):
                continue

        for period in sorted(by_period.keys()):
            rec = by_period[period]
            for series, store in series_areas.items():
                if series in rec:
                    store.append((period, area, rec[series]))

    if not series_areas[REFINERY_UTIL_SERIES["utilization_pct"]]:
        set(CACHE_KEY_UTIL, result, CACHE_TTL_REFINERY_UTIL,
            last_updated=datetime.now(timezone.utc).isoformat())
        return result

    all_periods = sorted({p for p, _, _ in
                          series_areas[REFINERY_UTIL_SERIES["utilization_pct"]]})
    result["dates"] = all_periods

    def _grid(series_key, area):
        lookup = {p: v for p, a, v in series_areas[series_key] if a == area}
        return [lookup.get(p) for p in all_periods]

    for area in areas:
        result["utilization_pct"][area] = _grid(
            REFINERY_UTIL_SERIES["utilization_pct"], area)
        result["crude_inputs"][area] = _grid(
            REFINERY_UTIL_SERIES["crude_inputs"], area)
        result["operable_capacity"][area] = _grid(
            REFINERY_UTIL_SERIES["operable_capacity"], area)

    result["available"] = True
    set(CACHE_KEY_UTIL, result, CACHE_TTL_REFINERY_UTIL,
        last_updated=datetime.now(timezone.utc).isoformat())
    return result
